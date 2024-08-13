import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from paraview.simple import *
from paraview.simple import SaveScreenshot

def load_data(file_path):
    """Load the data into ParaView."""
    reader = OpenDataFile(file_path)
    return reader

def slice_data(data, normal=[1, 0, 0], origin=[0, 0, 0]):
    """Create a slice of the data."""
    slice_filter = Slice(Input=data)
    slice_filter.SliceType.Normal = normal
    slice_filter.SliceType.Origin = origin
    return slice_filter

def display_scalar(data, scalar_name):
    """Display a scalar field (e.g., Cp or Mach number)."""
    data_display = Show(data)
    data_display.ColorArrayName = [None, scalar_name]
    return data_display

def take_screenshot(filename, view=None):
    """Save a screenshot of the current view."""
    if view is None:
        view = GetActiveView()
    SaveScreenshot(filename, view=view, quality=100)

def save_to_excel_with_openpyxl(image_paths, excel_file):
    """Save image paths to an Excel file using openpyxl."""
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image
    
    workbook = Workbook()
    
    for sheet_name, paths in image_paths.items():
        worksheet = workbook.create_sheet(title=sheet_name)
        
        for idx, image_info in enumerate(paths):
            label = image_info["Label"]
            image_path = image_info["Path"]
            
            worksheet.cell(row=idx + 2, column=1, value=label)
            
            img = Image(image_path)
            img.anchor = f'A{idx + 2}'
            worksheet.add_image(img)
    
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])
    
    workbook.save(excel_file)

def main():
    input_file = "path/to/your/file.vtk"
    output_excel = "output_screenshots.xlsx"
    
    # Load Data
    data = load_data(input_file)
    
    # Create output directory if not exists
    output_dir = "screenshots"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Store image paths
    image_paths = {
        "slices": [],
        "cp": [],
        "mach": []
    }
    
    # Example Operations
    # 1. Slice
    slice_filter = slice_data(data)
    Show(slice_filter)
    slice_img = os.path.join(output_dir, "slice.png")
    take_screenshot(slice_img)
    image_paths["slices"].append({"Label": "Slice", "Path": slice_img})
    
    # 2. Display Cp
    cp_display = display_scalar(data, "Cp")
    cp_img = os.path.join(output_dir, "cp.png")
    take_screenshot(cp_img)
    image_paths["cp"].append({"Label": "Cp", "Path": cp_img})
    
    # 3. Display Mach
    mach_display = display_scalar(data, "Mach")
    mach_img = os.path.join(output_dir, "mach.png")
    take_screenshot(mach_img)
    image_paths["mach"].append({"Label": "Mach", "Path": mach_img})
    
    # Save to Excel with xlsxwriter or openpyxl
    save_to_excel_with_openpyxl(image_paths, output_excel)

if __name__ == "__main__":
    main()
